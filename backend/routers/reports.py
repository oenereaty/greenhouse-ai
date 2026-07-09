"""리포트 탭 — 환경데이터 통계·시각화 + 농작업 기록 + 병해 로그, 주기 자동 생성."""
import io

import openpyxl
from fastapi import APIRouter, HTTPException
from pydantic import BaseModel
from starlette.responses import StreamingResponse

from backend import scheduler
from tools import report_generator, report_store

router = APIRouter(prefix="/api/reports", tags=["reports"])


class ReportSettingsIn(BaseModel):
    enabled: bool
    interval_days: int


@router.get("/settings")
def get_settings() -> dict:
    return scheduler.load_report_settings()


@router.put("/settings")
def put_settings(body: ReportSettingsIn) -> dict:
    if body.interval_days < 1:
        body.interval_days = 7
    return scheduler.save_report_settings(body.enabled, body.interval_days)


@router.get("/status")
def get_status() -> dict:
    return scheduler.get_report_status()


class GenerateIn(BaseModel):
    days: int = 7


@router.post("/generate")
def generate(body: GenerateIn) -> dict:
    days = body.days if body.days >= 1 else 7
    report = report_generator.build_report(days=days)
    report_id = report_store.save(report)
    return {"report_id": report_id, **report}


@router.get("/list")
def list_reports() -> list[dict]:
    return report_store.list_reports()


@router.get("/{report_id}")
def get_report(report_id: str) -> dict:
    report = report_store.load(report_id)
    if report is None:
        raise HTTPException(404, "리포트를 찾을 수 없습니다.")
    return {"report_id": report_id, **report}


@router.get("/{report_id}/export")
def export_report(report_id: str) -> StreamingResponse:
    report = report_store.load(report_id)
    if report is None:
        raise HTTPException(404, "리포트를 찾을 수 없습니다.")

    wb = openpyxl.Workbook()

    ws_env = wb.active
    ws_env.title = "환경데이터 통계"
    ws_env.append(["날짜", "평균온도", "최저온도", "최고온도", "평균습도", "평균CO2", "평균일사량", "측정건수"])
    for d in report["env"]["daily"]:
        ws_env.append([
            d["date"], d["avg_temp"], d["min_temp"], d["max_temp"],
            d["avg_rh"], d["avg_co2"], d["avg_solar"], d["reading_count"],
        ])

    ws_diary = wb.create_sheet("농작업 기록")
    ws_diary.append(["날짜", "시각", "내용", "태그", "약제"])
    for e in report["diary"]:
        ws_diary.append([
            e["date"], e.get("time", ""), e.get("content", ""),
            ", ".join(e.get("tags", [])), ", ".join(e.get("pesticides", [])),
        ])

    ws_disease = wb.create_sheet("병해 로그")
    ws_disease.append(["날짜", "내용", "태그", "약제"])
    for e in report["disease_log"]:
        ws_disease.append([
            e["date"], e.get("content", ""),
            ", ".join(e.get("tags", [])), ", ".join(e.get("pesticides", [])),
        ])

    buf = io.BytesIO()
    wb.save(buf)
    buf.seek(0)
    period = report["period"]
    filename = f"report_{period['start']}_{period['end']}.xlsx"
    return StreamingResponse(
        buf,
        media_type="application/vnd.openxmlformats-officedocument.spreadsheetml.sheet",
        headers={"Content-Disposition": f"attachment; filename={filename}"},
    )
